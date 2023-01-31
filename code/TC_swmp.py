# 爬運動中心
import time

from datetime import datetime
import openpyxl
import requests
import time
import os
import re

from bs4 import BeautifulSoup



URL = "https://tccsc.cyc.org.tw/api"
output_dir = '../file'
file_name = 'totalpeople.xlsx'
# * 表示比對內容可以出現 0 次或多次
# ? 表示比對內容可以出現 0 或 1 次
# + 表示比對內容至少出現 1 次，沒有上限
# \ 則是可以粗分成兩種來說明：

if __name__ == "__main__":

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    if not os.path.isfile(f'{output_dir}/{file_name}'):
        os.chdir(output_dir)
        wb = openpyxl.Workbook()  # 建立空白的 Excel 活頁簿物件
        # wb.save(file_name)  # 儲存檔案
        print("建立檔案")
    else:
        os.chdir(output_dir)
        print("開啟檔案")
        wb = openpyxl.load_workbook(file_name, data_only=True)  # 開啟現有的 Excel 活頁簿物件
        # wb.save('new.xlsx')  # 儲存檔案
    #print((str(datetime.now())[:19]))
    s1 = wb['Sheet']  # 開啟工作表 1
    cnt = int(s1.cell(1, 4).value)
    while(1):
        response = requests.get(url=URL)
        #print(response)
        soup = BeautifulSoup(response.text, 'html.parser')
        k = soup.text.split(',')
        # res = soup.find("body", classes_="vsc-initialized")
        # print(res)
        swim_peop, gym_peop = re.findall(r"\d+", k[3]), re.findall(r"\d+", k[0])
        # print(swim_peop, gym_peop)
        # print(k)

        print(cnt)
        s1['A1'].value = 'TIME'  # 儲存格 A1 內容為 apple
        s1['B1'].value = 'GYM'  # 儲存格 A2 內容為 orange
        s1['C1'].value = 'SWIM'  # 儲存格 A3 內容為 banana
        s1.cell(cnt, 8).value = (str(datetime.now())[:16])  # 儲存格 B1 內容 ( row=1, column=2 ) 為 100
        s1.cell(cnt, 9).value = str(gym_peop[0])  # 儲存格 B2 內容 ( row=2, column=2 ) 為 200
        s1.cell(cnt, 10).value = str(swim_peop[0])  # 儲存格 B3 內容 ( row=3, column=2 ) 為 300
        cnt += 1
        s1.cell(1, 4).value = str(cnt)
        wb.save(file_name)


        time.sleep(300)



